import csv
import io
from datetime import date

from flask import Blueprint, request, send_file
from openpyxl import Workbook, load_workbook

from ..extensions import db
from ..models import Schedule, Student
from ..routes.results import generate_ranking_results


import_export_bp = Blueprint("import_export", __name__)

REQUIRED_FIELDS = ["student_id", "name", "nationality", "department", "email"]
REQUIRED_DOCUMENTS = ["護照", "歷年成績單", "語言能力證明"]
GPA_THRESHOLD = 2.5
LANGUAGE_THRESHOLD = 60
IMPORT_FIELD_ALIASES = {
    "student_id": ["student_id", "student_code", "applicant_id", "application_id", "申請編號"],
    "name": ["name", "student_name", "applicant_name", "applicant", "姓名", "申請人姓名"],
    "nationality": ["nationality", "country", "國籍"],
    "department": ["department", "program", "major", "申請系所", "系所"],
    "email": ["email", "e_mail", "電子郵件"],
    "documents": ["documents", "document_status", "files", "文件", "文件狀態"],
    "gpa": ["gpa"],
    "language_score": ["language_score", "language", "語言成績"],
    "academic_score": ["academic_score", "academic", "學業成績"],
    "interview_score": ["interview_score", "interview", "面試成績"],
    "motivation_score": ["motivation_score", "motivation", "學習動機"],
    "bonus_score": ["bonus_score", "bonus", "加分"],
    "comment": ["comment", "note", "備註", "審查意見"],
}


def normalize_header(header):
    return (header or "").strip().lower().replace(" ", "_")


def first_row_value(row, field, default=""):
    for key in IMPORT_FIELD_ALIASES.get(field, [field]):
        value = row.get(normalize_header(key))
        if value not in (None, ""):
            return value

    return default


def normalize_import_row(row):
    normalized = {
        "student_id": str(first_row_value(row, "student_id")).strip(),
        "name": str(first_row_value(row, "name")).strip(),
        "nationality": str(first_row_value(row, "nationality")).strip(),
        "department": str(first_row_value(row, "department")).strip(),
        "email": str(first_row_value(row, "email")).strip(),
        "documents": parse_documents(first_row_value(row, "documents", [])),
        "gpa": to_float(first_row_value(row, "gpa")),
        "language_score": to_float(first_row_value(row, "language_score")),
        "academic_score": to_float(first_row_value(row, "academic_score")),
        "interview_score": to_float(first_row_value(row, "interview_score")),
        "motivation_score": to_float(first_row_value(row, "motivation_score")),
        "bonus_score": to_float(first_row_value(row, "bonus_score")),
        "comment": str(first_row_value(row, "comment")).strip(),
    }
    normalized["_raw"] = row
    return normalized


def read_rows_from_upload(storage):
    filename = (storage.filename or "").lower()

    if filename.endswith(".csv"):
        text = storage.stream.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [{normalize_header(key): value for key, value in row.items()} for row in reader]

    if filename.endswith(".xlsx"):
        workbook = load_workbook(storage, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [normalize_header(cell) for cell in rows[0]]
        parsed = []

        for row in rows[1:]:
            parsed.append({
                headers[index]: row[index] if index < len(row) else None
                for index in range(len(headers))
                if headers[index]
            })

        return parsed

    raise ValueError("僅支援 .xlsx 與 .csv 檔案。")


def parse_documents(raw_value):
    if raw_value is None:
        return []

    if isinstance(raw_value, list):
        return [str(item).strip() for item in raw_value if str(item).strip()]

    return [item.strip() for item in str(raw_value).split(",") if item.strip()]


def to_float(raw_value, default=0):
    if raw_value in (None, ""):
        return default

    try:
        return float(raw_value)
    except (TypeError, ValueError):
        return default


def has_scores(student):
    return any([
        student.academic_score,
        student.language_score,
        student.interview_score,
        student.motivation_score,
        student.bonus_score,
    ])


def classify_student(documents, gpa, language_score):
    return "可安排面試", "基本資料完整"


def upsert_student_from_import(row):
    normalized = normalize_import_row(row)
    raw_row = normalized["_raw"]
    student_code = normalized["student_id"]
    name = normalized["name"]
    nationality = normalized["nationality"]
    department = normalized["department"]
    email = normalized["email"]

    required_values = {
        "student_id": student_code,
        "name": name,
        "nationality": nationality,
        "department": department,
        "email": email,
    }
    missing = [key for key, value in required_values.items() if not value]
    if missing:
        return None, {
            **{key: value for key, value in normalized.items() if key != "_raw"},
            "student_id": student_code or "未提供",
            "status": "error",
            "message": f"缺少必要欄位: {', '.join(missing)}",
        }

    documents = normalized["documents"]
    gpa = normalized["gpa"]
    language_score = normalized["language_score"]
    academic_score = normalized["academic_score"] or (gpa * 25 if gpa else 0)
    interview_score = normalized["interview_score"]
    motivation_score = normalized["motivation_score"]
    bonus_score = normalized["bonus_score"]
    admission_status, document_status = classify_student(documents, gpa, language_score)

    student = Student.query.filter_by(student_code=student_code).first()
    creating = student is None

    if creating:
        student = Student(
            student_code=student_code,
            gender=str(raw_row.get("gender") or "Male"),
            age=int(to_float(raw_row.get("age"), 20)),
            grade=str(raw_row.get("grade") or "Year 1"),
            admission_type=str(raw_row.get("admission_type") or "International Student Admission"),
            application_date=date.today(),
            documents=[],
        )
        db.session.add(student)

    student.name = name
    student.chinese_name = str(raw_row.get("chinese_name") or "").strip() or None
    student.gender = str(raw_row.get("gender") or student.gender or "Male")
    student.age = int(to_float(raw_row.get("age"), student.age or 20))
    student.nationality = nationality
    student.department = department
    student.email = email
    student.phone = str(raw_row.get("phone") or "").strip() or None
    student.address = str(raw_row.get("address") or "").strip() or None
    student.passport_no = str(raw_row.get("passport_no") or "").strip() or None
    student.arc_no = str(raw_row.get("arc_no") or "").strip() or None
    student.grade = str(raw_row.get("grade") or student.grade or "Year 1")
    student.admission_type = str(raw_row.get("admission_type") or student.admission_type or "International Student Admission")
    student.documents = documents
    student.gpa = gpa
    student.academic_score = academic_score
    student.language_score = language_score
    student.interview_score = interview_score
    student.motivation_score = motivation_score
    student.bonus_score = bonus_score
    student.comment = normalized["comment"] or None
    student.document_status = document_status
    student.admission_status = admission_status
    student.note = str(raw_row.get("note") or student.note or "").strip() or None
    student.interview_status = student.interview_status or "未安排"

    return student, {
        "student_id": student_code,
        "name": name,
        "nationality": nationality,
        "department": department,
        "email": email,
        "documents": documents,
        "gpa": gpa,
        "language_score": language_score,
        "academic_score": academic_score,
        "interview_score": interview_score,
        "motivation_score": motivation_score,
        "bonus_score": bonus_score,
        "comment": normalized["comment"],
        "status": "created" if creating else "updated",
        "admission_status": admission_status,
        "document_status": document_status,
    }


@import_export_bp.post("/import/students")
def import_students():
    file = request.files.get("file")
    if file is None or not file.filename:
        return {"error": "請上傳 .xlsx 或 .csv 檔案。"}, 400

    try:
        rows = read_rows_from_upload(file)
    except ValueError as error:
        return {"error": str(error)}, 400

    imported = []
    errors = []

    for row in rows:
        student, result = upsert_student_from_import(row)
        if student is None:
            errors.append(result)
        else:
            imported.append(result)

    db.session.commit()

    return {
        "message": "匯入完成",
        "imported_count": len(imported),
        "error_count": len(errors),
        "preview": imported[:20],
        "errors": errors[:20],
    }


@import_export_bp.post("/students/auto-classify")
def auto_classify_students():
    updated = []

    for student in Student.query.order_by(Student.student_code.asc()).all():
        admission_status, document_status = classify_student(student.documents or [], student.gpa or 0, student.language_score or 0)
        student.admission_status = admission_status
        student.document_status = document_status
        updated.append({
            "student_id": student.student_code,
            "name": student.name,
            "nationality": student.nationality,
            "department": student.department,
            "email": student.email,
            "documents": student.documents or [],
            "gpa": student.gpa,
            "language_score": student.language_score,
            "academic_score": student.academic_score,
            "interview_score": student.interview_score,
            "motivation_score": student.motivation_score,
            "bonus_score": student.bonus_score,
            "comment": student.comment or "",
            "admission_status": admission_status,
            "document_status": document_status,
        })

    db.session.commit()
    return {"message": "分類完成", "updated_count": len(updated), "students": updated}


@import_export_bp.get("/export/results")
def export_results():
    results = generate_ranking_results()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Admission Results"
    sheet.append([
        "student_id",
        "name",
        "department",
        "total_score",
        "rank",
        "recommendation",
        "admission_status",
        "comment",
    ])

    for item in results:
        sheet.append([
            item["student_id"],
            item["name"],
            item["department"],
            item["total_score"],
            item["rank"],
            item["recommendation"],
            item["admission_status"],
            item["comment"],
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="admission_results.xlsx",
    )


@import_export_bp.get("/export/interviews")
def export_interviews():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Interview List"
    sheet.append([
        "student_id",
        "name",
        "department",
        "interview_date",
        "interview_time",
        "mode",
        "location",
        "teacher",
        "status",
    ])

    for student in Student.query.order_by(Student.interview_date.asc(), Student.student_code.asc()).all():
        if not student.interview_date:
            continue

        sheet.append([
            student.student_code,
            student.name,
            student.department,
            student.interview_date.isoformat(),
            "",
            student.interview_mode or "",
            student.interview_room or "",
            student.assigned_teacher or "",
            student.interview_status or "",
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="interview_list.xlsx",
    )


@import_export_bp.get("/export/interview-notifications")
def export_interview_notifications():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Interview Notifications"
    sheet.append([
        "申請編號",
        "申請人姓名",
        "Email",
        "面試日期",
        "面試時間",
        "面試方式",
        "面試地點 / 線上連結",
        "負責教師",
        "面試批次",
        "備註",
    ])

    schedules = Schedule.query.order_by(
        Schedule.event_date.asc(),
        Schedule.event_time.asc(),
        Schedule.title.asc(),
        Schedule.id.asc(),
    ).all()

    for schedule in schedules:
        student = schedule.student
        if student is None:
            continue

        sheet.append([
            student.student_code,
            student.name,
            student.email or "",
            schedule.event_date.isoformat(),
            schedule.event_time.strftime("%H:%M"),
            student.interview_mode or "",
            schedule.location or student.interview_room or "",
            student.assigned_teacher or "",
            schedule.title,
            student.note or "",
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="interview_notification_list.xlsx",
    )


@import_export_bp.get("/export/scoring-template")
def export_scoring_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Interview Scores"
    sheet.append([
        "student_id",
        "name",
        "department",
        "academic_score",
        "language_score",
        "interview_score",
        "motivation_score",
        "bonus_score",
        "comment",
    ])

    for student in Student.query.order_by(Student.student_code.asc()).all():
        sheet.append([
            student.student_code,
            student.name,
            student.department,
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="scoring_template.xlsx",
    )


@import_export_bp.post("/import/scores")
def import_scores():
    file = request.files.get("file")
    if file is None or not file.filename:
        return {"error": "請上傳已評分的 .xlsx 或 .csv 檔案。"}, 400

    try:
        rows = read_rows_from_upload(file)
    except ValueError as error:
        return {"error": str(error)}, 400

    updated = []
    errors = []

    for row in rows:
        student_code = str(first_row_value(row, "student_id")).strip()
        student = Student.query.filter_by(student_code=student_code).first()

        if not student:
            errors.append({"student_id": student_code or "未提供", "message": "找不到申請人"})
            continue

        student.academic_score = to_float(row.get("academic_score"))
        student.language_score = to_float(row.get("language_score"))
        student.interview_score = to_float(row.get("interview_score"))
        student.motivation_score = to_float(row.get("motivation_score"))
        student.bonus_score = to_float(row.get("bonus_score"))
        student.comment = str(row.get("comment") or "").strip() or student.comment
        student.total_score = 0
        student.ranking = None
        student.recommendation = None

        updated.append({
            "student_id": student.student_code,
            "name": student.name,
            "department": student.department,
            "academic_score": student.academic_score,
            "language_score": student.language_score,
            "interview_score": student.interview_score,
            "motivation_score": student.motivation_score,
            "bonus_score": student.bonus_score,
            "comment": student.comment or "",
            "status": "已匯入評分",
        })

    db.session.commit()

    return {
        "message": "評分匯入完成",
        "updated_count": len(updated),
        "error_count": len(errors),
        "preview": updated[:20],
        "errors": errors[:20],
    }
